"""Excel estate exporter — produces a comprehensive multi-sheet .xlsx workbook.

Sheets
------
1.  Summary              — High-level stats for every configured tool
2.  Signals & Metrics    — Every telemetry signal collected across all tools
3.  Services & Apps      — Applications / APIs sending telemetry data
4.  Scrape Targets       — Prometheus scrape jobs and instance health
5.  Alert Rules          — All alerts across all tools, sorted by severity
6.  Recording Rules & SLOs — SLOs, recording rules, and computation expressions
7.  Dashboards           — All dashboards with folder / tag / panel metadata
8.  Dashboard Panels     — Panel-level detail with queries / computations
9.  Datasources          — Data connections configured in the observability stack
10. Alert Receivers      — Notification channels and integration types
11. Tracing              — Distributed-trace services and sampled operations
12. Tool Topology        — How tools communicate with each other
13. OTel Pipelines       — OpenTelemetry Collector receivers, exporters, pipelines
14. Label Inventory      — Unique label/tag keys and where they appear
15. Extraction Log       — Success / error log from data collection
"""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from observascore.model import ObservabilityEstate

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_DARK = {
    "summary":     "1F4E79",
    "signals":     "0D5E5E",
    "services":    "1B4332",
    "targets":     "1C3A6E",
    "alerts":      "7B1C1C",
    "slos":        "4A1572",
    "dashboards":  "7D4000",
    "panels":      "5C3D00",
    "datasources": "005B5B",
    "receivers":   "5C001A",
    "tracing":     "003B4D",
    "topology":    "2E4057",
    "otel":        "3B0764",
    "labels":      "3D3B00",
    "log":         "1A1A1A",
}

_SEV = {
    "critical": "FF9999",
    "high":     "FFCC99",
    "medium":   "FFFB99",
    "low":      "CCE5FF",
    "info":     "E8E8E8",
}

_HEALTH = {
    "up":      "C6EFCE",
    "down":    "FFC7CE",
    "unknown": "FFFFCC",
}

_STATUS = {
    "adopted": "C6EFCE",
    "partial": "FFFFCC",
    "absent":  "FFC7CE",
    "yes":     "C6EFCE",
    "no":      "FFC7CE",
}

_WHITE = "FFFFFF"

_BOLD_WHITE = Font(bold=True, color=_WHITE, name="Calibri", size=10)
_DEFAULT    = Font(name="Calibri", size=10)
_HEADER_AL  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_AL    = Alignment(vertical="top", wrap_text=True)
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _sev_fill(sev: Optional[str]) -> Optional[str]:
    return _SEV.get(str(sev).lower()) if sev else None


def _labels_str(d: dict) -> str:
    return ", ".join(f"{k}={v}" for k, v in (d or {}).items())


def _truncate(val, n: int = 500) -> str:
    s = str(val) if val is not None else ""
    return s if len(s) <= n else s[:n] + "…"


def _yes_no(val) -> str:
    return "Yes" if val else "No"


# ---------------------------------------------------------------------------
# ExcelExporter
# ---------------------------------------------------------------------------

class ExcelExporter:
    """Converts an :class:`~observascore.model.ObservabilityEstate` into a
    multi-sheet Excel workbook — one sheet per observability concern."""

    MAX_ROWS = 5_000   # hard cap per sheet to avoid giant files

    def export(self, estate: ObservabilityEstate, output_dir: Path) -> Path:
        """Build the workbook and write it to *output_dir*.

        Returns the absolute path of the saved file.
        """
        wb = Workbook()
        # Remove the default "Sheet" that openpyxl creates
        for name in list(wb.sheetnames):
            del wb[name]

        self._sheet_summary(wb, estate)
        self._sheet_signals(wb, estate)
        self._sheet_services(wb, estate)
        self._sheet_scrape_targets(wb, estate)
        self._sheet_alert_rules(wb, estate)
        self._sheet_slos_recording(wb, estate)
        self._sheet_dashboards(wb, estate)
        self._sheet_dashboard_panels(wb, estate)
        self._sheet_datasources(wb, estate)
        self._sheet_alert_receivers(wb, estate)
        self._sheet_tracing(wb, estate)
        self._sheet_topology(wb, estate)
        self._sheet_otel(wb, estate)
        self._sheet_label_inventory(wb, estate)
        self._sheet_extraction_log(wb, estate)

        output_dir.mkdir(parents=True, exist_ok=True)
        safe = estate.client_name.lower().replace(" ", "-").replace("/", "-")
        ts   = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
        fname = f"observascore-export-{safe}-{estate.environment}-{ts}.xlsx"
        path  = output_dir / fname
        wb.save(str(path))
        logger.info("Excel export saved: %s", path)
        return path

    # -----------------------------------------------------------------------
    # Low-level helpers
    # -----------------------------------------------------------------------

    @staticmethod
    def _add_ws(wb: Workbook, title: str, tab_color: str):
        ws = wb.create_sheet(title=title)
        ws.sheet_properties.tabColor = tab_color
        return ws

    @staticmethod
    def _header(ws, headers: list[str], fill_key: str, row: int = 1) -> None:
        bg = _fill(_DARK.get(fill_key, "1F4E79"))
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font      = _BOLD_WHITE
            cell.fill      = bg
            cell.alignment = _HEADER_AL
            cell.border    = _BORDER

    @staticmethod
    def _row(ws, row_num: int, values: list, cell_fills: list[Optional[str]] = ()) -> None:
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=_truncate(val))
            cell.font      = _DEFAULT
            cell.alignment = _CELL_AL
            cell.border    = _BORDER
            idx = c - 1
            if idx < len(cell_fills) and cell_fills[idx]:
                cell.fill = _fill(cell_fills[idx])

    @staticmethod
    def _auto_width(ws, min_w: int = 8, max_w: int = 60) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            best = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[letter].width = max(min_w, min(best + 2, max_w))

    @staticmethod
    def _freeze(ws, cell: str = "A2") -> None:
        ws.freeze_panes = cell

    def _finalize(self, ws) -> None:
        self._auto_width(ws)
        self._freeze(ws)

    def _title_row(self, ws, text: str, fill_key: str, ncols: int) -> None:
        """Write a full-width title row above the header."""
        end = get_column_letter(ncols)
        ws.merge_cells(f"A1:{end}1")
        cell = ws["A1"]
        cell.value     = text
        cell.font      = Font(bold=True, size=12, color=_WHITE, name="Calibri")
        cell.fill      = _fill(_DARK.get(fill_key, "1F4E79"))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        # actual header on row 2
        ws.freeze_panes = "A3"

    def _note_truncated(self, ws, row: int, ncols: int) -> None:
        ws.cell(row=row, column=1,
                value=f"⚠ Results capped at {self.MAX_ROWS:,} rows — run with a scoped config for full data.")

    # -----------------------------------------------------------------------
    # Sheet 1 — Summary
    # -----------------------------------------------------------------------

    def _sheet_summary(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "📋 Summary", _DARK["summary"])
        s  = estate.summary
        configured = set(estate.configured_tools)

        ncols = 3
        self._title_row(ws,
            f"ObservaScore Estate Export  ·  {estate.client_name}  ·  {estate.environment}  ·  {estate.timestamp}",
            "summary", ncols)
        self._header(ws, ["Category", "Metric", "Value"], "summary", row=2)

        tool_map = {
            "Prometheus": "prometheus", "Grafana": "grafana", "Loki": "loki",
            "Jaeger": "jaeger", "AlertManager": "alertmanager", "Tempo": "tempo",
            "Elasticsearch": "elasticsearch", "OTel Collector": "otel_collector",
            "AppDynamics": "appdynamics", "Datadog": "datadog", "Dynatrace": "dynatrace",
        }

        rows = [
            # ── Client metadata ──
            ("Client", "Name",               estate.client_name),
            ("Client", "Environment",        estate.environment),
            ("Client", "Exported At",        estate.timestamp),
            ("Client", "Configured Tools",   ", ".join(configured) or "none"),
            ("Client", "Extraction Errors",  len(s.extraction_errors)),
            # ── Prometheus ──
            ("Prometheus", "Scrape Targets",      s.prometheus_targets),
            ("Prometheus", "Targets Up",           s.prometheus_targets_up),
            ("Prometheus", "Targets Down",         max(0, s.prometheus_targets - s.prometheus_targets_up)),
            ("Prometheus", "Alert Rules",          s.prometheus_alert_rules),
            ("Prometheus", "Recording Rules",      s.prometheus_recording_rules),
            ("Prometheus", "Metrics Sampled",      s.prometheus_metrics_sampled),
            # ── Grafana ──
            ("Grafana", "Dashboards",          s.grafana_dashboards),
            ("Grafana", "Folders",             s.grafana_folders),
            ("Grafana", "Datasources",         s.grafana_datasources),
            ("Grafana", "Alert Rules",         s.grafana_alert_rules),
            # ── Loki ──
            ("Loki", "Log Labels",             s.loki_labels),
            # ── Jaeger ──
            ("Jaeger", "Services Traced",      s.jaeger_services),
            ("Jaeger", "Total Operations",     s.jaeger_operations),
            # ── AlertManager ──
            ("AlertManager", "Receivers",               s.alertmanager_receivers),
            ("AlertManager", "Active Silences",         s.alertmanager_silences),
            ("AlertManager", "Integration Types",       ", ".join(s.alertmanager_integrations) or "none"),
            # ── Tempo ──
            ("Tempo", "Services Traced",       s.tempo_services),
            # ── Elasticsearch ──
            ("Elasticsearch", "Indices",             s.elasticsearch_indices),
            ("Elasticsearch", "Data Streams",        s.elasticsearch_data_streams),
            # ── OTel Collector ──
            ("OTel Collector", "Receivers",       ", ".join(s.otel_receivers) or "none"),
            ("OTel Collector", "Exporters",       ", ".join(s.otel_exporters) or "none"),
            ("OTel Collector", "Pipelines",       s.otel_pipelines),
            # ── AppDynamics ──
            ("AppDynamics", "Applications",              s.appdynamics_applications),
            ("AppDynamics", "Tiers / Agents",            s.appdynamics_tiers),
            ("AppDynamics", "Health Rules",              s.appdynamics_health_rules),
            ("AppDynamics", "Business Transactions",     s.appdynamics_business_transactions),
            ("AppDynamics", "EUM — End User Monitoring", _yes_no(s.appdynamics_has_eum)),
            ("AppDynamics", "SIM — Server Infrastructure", _yes_no(s.appdynamics_has_sim)),
            ("AppDynamics", "Database Monitoring",       _yes_no(s.appdynamics_has_db_monitoring)),
            ("AppDynamics", "Apps with Baselines",       s.appdynamics_apps_with_baselines),
            # ── Datadog ──
            ("Datadog", "Monitors",                    s.datadog_monitors),
            ("Datadog", "Monitors with Notifications", s.datadog_monitors_with_notifications),
            ("Datadog", "Dashboards",                  s.datadog_dashboards),
            ("Datadog", "Hosts",                       s.datadog_hosts),
            ("Datadog", "SLOs",                        s.datadog_slos),
            ("Datadog", "Synthetic Tests",             s.datadog_synthetics),
            ("Datadog", "APM",                         _yes_no(s.datadog_has_apm)),
            ("Datadog", "Log Management",              _yes_no(s.datadog_has_log_management)),
            ("Datadog", "Security Monitoring",         _yes_no(s.datadog_has_security_monitoring)),
            ("Datadog", "Service Catalog",             _yes_no(s.datadog_has_service_catalog)),
            # ── Dynatrace ──
            ("Dynatrace", "Services",                     s.dynatrace_services),
            ("Dynatrace", "Hosts",                        s.dynatrace_hosts),
            ("Dynatrace", "Applications",                 s.dynatrace_applications),
            ("Dynatrace", "Open Problems",                s.dynatrace_problems_open),
            ("Dynatrace", "SLOs",                         s.dynatrace_slos),
            ("Dynatrace", "Synthetic Monitors",           s.dynatrace_synthetics),
            ("Dynatrace", "Alerting Profiles",            s.dynatrace_alerting_profiles),
            ("Dynatrace", "Notification Integrations",    s.dynatrace_notification_integrations),
            ("Dynatrace", "Log Management",               _yes_no(s.dynatrace_has_log_management)),
            ("Dynatrace", "Real User Monitoring (RUM)",   _yes_no(s.dynatrace_has_rum)),
        ]

        for i, (cat, metric, val) in enumerate(rows, start=3):
            fill_c = None
            if isinstance(val, str) and val in ("Yes", "No"):
                fill_c = _STATUS.get(val.lower())
            elif cat != "Client" and tool_map.get(cat) not in configured:
                fill_c = "F0F0F0"   # grey out rows for tools not in this run
            self._row(ws, i, [cat, metric, val], [None, None, fill_c])

        self._auto_width(ws)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 42

    # -----------------------------------------------------------------------
    # Sheet 2 — Signals & Metrics
    # -----------------------------------------------------------------------

    def _sheet_signals(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "📡 Signals & Metrics", _DARK["signals"])
        headers = ["#", "Identifier / Metric Name", "Signal Type", "Semantic Type",
                   "Source Tool", "Cardinality Estimate", "Labels"]
        self._title_row(ws, "Signals & Metrics — all telemetry signals collected across tools", "signals", len(headers))
        self._header(ws, headers, "signals", row=2)

        signals = estate.signals[: self.MAX_ROWS]
        for i, sig in enumerate(signals, 1):
            self._row(ws, i + 2, [
                i,
                sig.identifier,
                sig.signal_type.value if hasattr(sig.signal_type, "value") else str(sig.signal_type),
                sig.semantic_type or "",
                sig.source_tool,
                sig.cardinality_estimate if sig.cardinality_estimate is not None else "",
                _labels_str(sig.labels),
            ])

        if len(estate.signals) > self.MAX_ROWS:
            self._note_truncated(ws, len(signals) + 4, len(headers))

        self._finalize(ws)
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["G"].width = 50

    # -----------------------------------------------------------------------
    # Sheet 3 — Services & Applications
    # -----------------------------------------------------------------------

    def _sheet_services(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🎯 Services & Apps", _DARK["services"])
        headers = ["#", "Service / Application Name", "Source Tool", "Tier",
                   "Owner", "Operation Count", "Sample Operations (first 15)"]
        self._title_row(ws, "Services & Applications — APIs and services instrumented with observability agents", "services", len(headers))
        self._header(ws, headers, "services", row=2)

        svcs = estate.services[: self.MAX_ROWS]
        for i, svc in enumerate(svcs, 1):
            ops_sample = "; ".join(svc.operations[:15]) if svc.operations else ""
            self._row(ws, i + 2, [
                i,
                svc.name,
                svc.source_tool,
                svc.tier or "",
                svc.owner or "",
                len(svc.operations),
                ops_sample,
            ])

        self._finalize(ws)
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["G"].width = 60

    # -----------------------------------------------------------------------
    # Sheet 4 — Scrape Targets
    # -----------------------------------------------------------------------

    def _sheet_scrape_targets(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🔍 Scrape Targets", _DARK["targets"])
        headers = ["#", "Job", "Instance", "Health", "Last Scrape Error", "Labels", "Source Tool"]
        self._title_row(ws, "Prometheus Scrape Targets — what Prometheus is monitoring", "targets", len(headers))
        self._header(ws, headers, "targets", row=2)

        targets = estate.scrape_targets[: self.MAX_ROWS]
        for i, t in enumerate(targets, 1):
            health_fill = _HEALTH.get(str(t.health).lower())
            self._row(ws, i + 2, [
                i, t.job, t.instance, t.health,
                t.last_scrape_error or "",
                _labels_str(t.labels),
                t.source_tool,
            ], [None, None, None, health_fill])

        self._finalize(ws)
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["F"].width = 50

    # -----------------------------------------------------------------------
    # Sheet 5 — Alert Rules
    # -----------------------------------------------------------------------

    def _sheet_alert_rules(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🚨 Alert Rules", _DARK["alerts"])
        headers = ["#", "Alert / Monitor Name", "Severity", "Classification",
                   "Source Tool", "Group", "Expression / Condition",
                   "Duration", "Labels", "Summary / Description", "Runbook URL"]
        self._title_row(ws, "Alert Rules — all alerts and monitors configured across every tool", "alerts", len(headers))
        self._header(ws, headers, "alerts", row=2)

        sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
        rules = sorted(estate.alert_rules, key=lambda r: sev_order.get(str(r.severity).lower(), 9))
        rules = rules[: self.MAX_ROWS]

        for i, r in enumerate(rules, 1):
            sev  = str(r.severity).lower() if r.severity else ""
            clf  = r.classification.value if hasattr(r.classification, "value") else str(r.classification)
            self._row(ws, i + 2, [
                i,
                r.name,
                r.severity or "",
                clf,
                r.source_tool,
                r.group or "",
                r.expression,
                r.for_duration or "",
                _labels_str(r.labels),
                r.annotations.get("summary", r.annotations.get("description", "")) if r.annotations else "",
                r.runbook_url or "",
            ], [None, None, _sev_fill(sev)])

        self._finalize(ws)
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["G"].width = 55
        ws.column_dimensions["K"].width = 40

    # -----------------------------------------------------------------------
    # Sheet 6 — Recording Rules & SLOs
    # -----------------------------------------------------------------------

    def _sheet_slos_recording(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "📊 Recording Rules & SLOs", _DARK["slos"])
        headers = ["#", "Name", "Type", "Source Tool", "Group", "Expression / Target", "Labels"]
        self._title_row(ws, "Recording Rules & SLOs — computed metrics, burn rates, and service-level objectives", "slos", len(headers))
        self._header(ws, headers, "slos", row=2)

        rr = estate.recording_rules[: self.MAX_ROWS]
        for i, r in enumerate(rr, 1):
            self._row(ws, i + 2, [
                i, r.name, "Recording Rule", r.source_tool,
                r.group or "", r.expression, _labels_str(r.labels),
            ])

        self._finalize(ws)
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["F"].width = 55

    # -----------------------------------------------------------------------
    # Sheet 7 — Dashboards
    # -----------------------------------------------------------------------

    def _sheet_dashboards(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "📈 Dashboards", _DARK["dashboards"])
        headers = ["#", "Dashboard Title", "Source Tool", "Folder", "Tags",
                   "Panel Count", "Has Variables", "Last Modified", "Owner", "UID"]
        self._title_row(ws, "Dashboards — all dashboards available across every observability tool", "dashboards", len(headers))
        self._header(ws, headers, "dashboards", row=2)

        dboards = estate.dashboards[: self.MAX_ROWS]
        for i, d in enumerate(dboards, 1):
            self._row(ws, i + 2, [
                i,
                d.title,
                d.source_tool,
                d.folder or "",
                ", ".join(d.tags) if d.tags else "",
                len(d.panels),
                _yes_no(d.has_templating),
                d.last_modified or "",
                d.owner or "",
                d.uid,
            ])

        self._finalize(ws)
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["E"].width = 30

    # -----------------------------------------------------------------------
    # Sheet 8 — Dashboard Panels
    # -----------------------------------------------------------------------

    def _sheet_dashboard_panels(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🖼 Dashboard Panels", _DARK["panels"])
        headers = ["#", "Dashboard Title", "Source Tool", "Folder",
                   "Panel Title", "Panel Type", "Unit", "Has Thresholds",
                   "Queries / Computations"]
        self._title_row(ws, "Dashboard Panels — individual panels with their queries and computations", "panels", len(headers))
        self._header(ws, headers, "panels", row=2)

        row_num = 3
        total   = 0
        for d in estate.dashboards:
            for panel in d.panels:
                if total >= self.MAX_ROWS:
                    break
                queries_str = " | ".join(panel.queries) if panel.queries else ""
                self._row(ws, row_num, [
                    total + 1,
                    d.title,
                    d.source_tool,
                    d.folder or "",
                    panel.title,
                    panel.panel_type,
                    panel.unit or "",
                    _yes_no(panel.has_thresholds),
                    queries_str,
                ])
                row_num += 1
                total   += 1
            if total >= self.MAX_ROWS:
                break

        if total == 0:
            ws.cell(row=3, column=1, value="No panel data available — Grafana panels require api_key with dashboard read access.")

        if total >= self.MAX_ROWS:
            self._note_truncated(ws, row_num + 1, len(headers))

        self._finalize(ws)
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["E"].width = 38
        ws.column_dimensions["I"].width = 60

    # -----------------------------------------------------------------------
    # Sheet 9 — Datasources
    # -----------------------------------------------------------------------

    def _sheet_datasources(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🔌 Datasources", _DARK["datasources"])
        headers = ["#", "Name", "Type", "URL", "Configured In", "Is Default", "Reachable"]
        self._title_row(ws, "Datasources — data connections configured in the observability stack", "datasources", len(headers))
        self._header(ws, headers, "datasources", row=2)

        for i, ds in enumerate(estate.datasources, 1):
            r_str   = "Yes" if ds.reachable else ("No" if ds.reachable is False else "Unknown")
            r_fill  = _STATUS.get(r_str.lower())
            def_str = _yes_no(ds.is_default)
            self._row(ws, i + 2, [
                i, ds.name, ds.ds_type, ds.url or "", ds.source_tool, def_str, r_str,
            ], [None, None, None, None, None, None, r_fill])

        self._finalize(ws)
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["D"].width = 40

    # -----------------------------------------------------------------------
    # Sheet 10 — Alert Receivers
    # -----------------------------------------------------------------------

    def _sheet_alert_receivers(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "📣 Alert Receivers", _DARK["receivers"])
        headers = ["#", "Receiver Name", "Integration Types", "Source Tool", "Notes"]
        self._title_row(ws, "Alert Receivers & Notification Channels — how alerts reach the on-call team", "receivers", len(headers))
        self._header(ws, headers, "receivers", row=2)

        for i, r in enumerate(estate.alert_receivers, 1):
            types = ", ".join(r.receiver_types) if r.receiver_types else "unknown"
            # Compute a human-readable notes string from known integrations
            notes_parts = []
            if "pagerduty" in types.lower():
                notes_parts.append("PagerDuty on-call routing")
            if "slack" in types.lower():
                notes_parts.append("Slack notifications")
            if "email" in types.lower():
                notes_parts.append("Email alerts")
            if "opsgenie" in types.lower():
                notes_parts.append("OpsGenie escalation")
            if "webhook" in types.lower():
                notes_parts.append("Generic webhook")
            self._row(ws, i + 2, [
                i, r.name, types, "alertmanager", "; ".join(notes_parts),
            ])

        if not estate.alert_receivers:
            ws.cell(row=3, column=1, value="No alert receivers found — enable alertmanager in config to populate this sheet.")

        self._finalize(ws)
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 36

    # -----------------------------------------------------------------------
    # Sheet 11 — Tracing
    # -----------------------------------------------------------------------

    def _sheet_tracing(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🌐 Tracing", _DARK["tracing"])
        headers = ["#", "Service Name", "Tracing Backend", "Tier",
                   "Operation Count", "Operations (first 20)"]
        self._title_row(ws, "Distributed Tracing — services instrumented for request tracing", "tracing", len(headers))
        self._header(ws, headers, "tracing", row=2)

        tracing_tools = {"jaeger", "tempo"}
        trace_svcs = [s for s in estate.services if s.source_tool in tracing_tools]
        trace_svcs = trace_svcs[: self.MAX_ROWS]

        for i, svc in enumerate(trace_svcs, 1):
            ops = "; ".join(svc.operations[:20]) if svc.operations else ""
            self._row(ws, i + 2, [
                i, svc.name, svc.source_tool, svc.tier or "",
                len(svc.operations), ops,
            ])

        if not trace_svcs:
            ws.cell(row=3, column=1,
                    value="No tracing data — enable jaeger or tempo in config to populate this sheet.")

        self._finalize(ws)
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["F"].width = 60

    # -----------------------------------------------------------------------
    # Sheet 12 — Tool Topology
    # -----------------------------------------------------------------------

    def _sheet_topology(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🗺 Tool Topology", _DARK["topology"])
        headers = ["#", "From (Source)", "To (Destination)", "Integration Type",
                   "Protocol", "Data Flow", "Endpoint / Details", "Notes"]
        self._title_row(ws, "Tool Topology — how observability tools communicate with each other", "topology", len(headers))
        self._header(ws, headers, "topology", row=2)

        tools   = set(estate.configured_tools)
        topo    = []

        # 1. Datasource connections (direct evidence from Grafana)
        for ds in estate.datasources:
            topo.append((
                ds.source_tool.capitalize(),
                ds.ds_type.capitalize(),
                "Datasource Query",
                "HTTP",
                "Metrics / Logs / Traces → Query",
                ds.url or "",
                f"{'[default] ' if ds.is_default else ''}Grafana datasource: {ds.name}",
            ))

        # 2. Prometheus → AlertManager  (standard Prometheus integration)
        if "prometheus" in tools and "alertmanager" in tools:
            topo.append((
                "Prometheus", "AlertManager",
                "Alert Routing", "HTTP",
                "Firing alerts → routing",
                "", "Prometheus evaluates rules and sends alerts to AlertManager for routing/silencing",
            ))

        # 3. OTel Collector ingest: external sources → OTel Collector
        for recv in estate.summary.otel_receivers:
            topo.append((
                recv.capitalize(), "OTel Collector",
                "Telemetry Ingest", "gRPC / HTTP",
                "Metrics / Traces / Logs → Collector",
                "", f"OTel receiver: {recv}",
            ))

        # 4. OTel Collector export: OTel Collector → backends
        for exp in estate.summary.otel_exporters:
            topo.append((
                "OTel Collector", exp.capitalize(),
                "Telemetry Export", "gRPC / HTTP",
                "Metrics / Traces / Logs → Backend",
                "", f"OTel exporter: {exp}",
            ))

        # 5. Commercial APM tools instrument application code
        for tool in ["appdynamics", "datadog", "dynatrace"]:
            if tool in tools:
                topo.append((
                    "Application Code",
                    tool.capitalize(),
                    "APM Agent Instrumentation",
                    "Agent (proprietary)",
                    "Metrics / Traces / Logs → APM Backend",
                    "",
                    f"{tool.capitalize()} agent installed in application runtime",
                ))

        # 6. Scrape topology: Prometheus ← scrape targets
        jobs = sorted({t.job for t in estate.scrape_targets})
        if jobs:
            topo.append((
                "Prometheus", "Scrape Targets",
                "Metrics Scrape", "HTTP (pull)",
                "Metrics ← /metrics endpoint",
                f"{len(jobs)} jobs: {', '.join(jobs[:8])}{'…' if len(jobs) > 8 else ''}",
                "Prometheus scrapes metrics from instrumented services and exporters",
            ))

        # 7. Loki ← log shippers
        if "loki" in tools:
            topo.append((
                "Log Shippers (Promtail / Alloy / Fluentd)", "Loki",
                "Log Ingest", "HTTP / gRPC",
                "Log streams → storage",
                "", "Applications and node-level log shippers push logs to Loki",
            ))

        # 8. Elasticsearch ← log shippers
        if "elasticsearch" in tools:
            topo.append((
                "Log Shippers (Logstash / Filebeat / Fluentd)", "Elasticsearch",
                "Log Ingest", "HTTP",
                "Log events → index",
                "", "Beats or Logstash pipelines ship logs to Elasticsearch",
            ))

        # 9. AlertManager → notification channels
        for recv in estate.alert_receivers:
            for rt in (recv.receiver_types or []):
                topo.append((
                    "AlertManager",
                    rt.capitalize(),
                    "Alert Notification",
                    "HTTP webhook / API",
                    "Alert events → notification",
                    "", f"AlertManager receiver: {recv.name}",
                ))

        for i, row in enumerate(topo, 1):
            self._row(ws, i + 2, [i] + list(row))

        self._finalize(ws)
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["G"].width = 42
        ws.column_dimensions["H"].width = 55

    # -----------------------------------------------------------------------
    # Sheet 13 — OTel Pipelines
    # -----------------------------------------------------------------------

    def _sheet_otel(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "⚙ OTel Pipelines", _DARK["otel"])
        headers = ["#", "Component Role", "Component Type", "Component Name",
                   "Telemetry Type Hint", "Notes"]
        self._title_row(ws, "OpenTelemetry Collector — receivers, processors, exporters and pipelines", "otel", len(headers))
        self._header(ws, headers, "otel", row=2)

        s = estate.summary
        otel_rows: list[tuple] = []

        for r in s.otel_receivers:
            ttype = _otel_telemetry_hint(r)
            otel_rows.append(("Input", "Receiver", r, ttype,
                              "Ingests telemetry from this source"))

        for e in s.otel_exporters:
            ttype = _otel_telemetry_hint(e)
            otel_rows.append(("Output", "Exporter", e, ttype,
                              "Forwards telemetry to this destination"))

        if s.otel_pipelines:
            otel_rows.append(("Config", "Pipeline Count",
                              str(s.otel_pipelines), "mixed",
                              "Total active pipelines in the collector config"))

        for i, r in enumerate(otel_rows, 1):
            self._row(ws, i + 2, [i] + list(r))

        if not otel_rows:
            ws.cell(row=3, column=1,
                    value="No OTel Collector data — enable otel_collector in config to populate this sheet.")

        self._finalize(ws)
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 32
        ws.column_dimensions["E"].width = 24
        ws.column_dimensions["F"].width = 45

    # -----------------------------------------------------------------------
    # Sheet 14 — Label Inventory
    # -----------------------------------------------------------------------

    def _sheet_label_inventory(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "🏷 Label Inventory", _DARK["labels"])
        headers = ["#", "Label / Tag Key", "Seen In Tools", "Contexts",
                   "Unique Value Count", "Sample Values (first 8)"]
        self._title_row(ws, "Label Inventory — all label/tag keys across metrics, alerts, and targets", "labels", len(headers))
        self._header(ws, headers, "labels", row=2)

        # Aggregate: key → {tools, contexts, values}
        inventory: dict[str, dict] = defaultdict(
            lambda: {"tools": set(), "contexts": set(), "values": set()}
        )

        for sig in estate.signals:
            for k, v in (sig.labels or {}).items():
                inventory[k]["tools"].add(sig.source_tool)
                inventory[k]["contexts"].add("signal")
                inventory[k]["values"].add(str(v)[:80])

        for ar in estate.alert_rules:
            for k, v in (ar.labels or {}).items():
                inventory[k]["tools"].add(ar.source_tool)
                inventory[k]["contexts"].add("alert_rule")
                inventory[k]["values"].add(str(v)[:80])

        for rr in estate.recording_rules:
            for k, v in (rr.labels or {}).items():
                inventory[k]["tools"].add(rr.source_tool)
                inventory[k]["contexts"].add("recording_rule")
                inventory[k]["values"].add(str(v)[:80])

        for st in estate.scrape_targets:
            for k, v in (st.labels or {}).items():
                inventory[k]["tools"].add(st.source_tool)
                inventory[k]["contexts"].add("scrape_target")
                inventory[k]["values"].add(str(v)[:80])

        sorted_keys = sorted(inventory.keys())[:self.MAX_ROWS]
        for i, key in enumerate(sorted_keys, 1):
            entry = inventory[key]
            self._row(ws, i + 2, [
                i,
                key,
                ", ".join(sorted(entry["tools"])),
                ", ".join(sorted(entry["contexts"])),
                len(entry["values"]),
                ", ".join(sorted(entry["values"])[:8]),
            ])

        if not inventory:
            ws.cell(row=3, column=1,
                    value="No label data collected — labels are populated when signals or alert rules are extracted.")

        self._finalize(ws)
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["F"].width = 60

    # -----------------------------------------------------------------------
    # Sheet 15 — Extraction Log
    # -----------------------------------------------------------------------

    def _sheet_extraction_log(self, wb: Workbook, estate: ObservabilityEstate) -> None:
        ws = self._add_ws(wb, "📝 Extraction Log", _DARK["log"])
        headers = ["#", "Source Tool", "Status", "Detail / Error Message"]
        self._title_row(ws, "Extraction Log — data collection results, errors, and warnings", "log", len(headers))
        self._header(ws, headers, "log", row=2)

        row_num = 3
        counter = 1

        # One success row per configured tool (errors listed separately)
        error_by_tool: dict[str, list[str]] = defaultdict(list)
        for err in estate.summary.extraction_errors:
            parts = err.split(": ", 1)
            tool  = parts[0] if len(parts) == 2 else "unknown"
            msg   = parts[1] if len(parts) == 2 else err
            error_by_tool[tool].append(msg)

        for tool in estate.configured_tools:
            errs = error_by_tool.get(tool, [])
            if errs:
                for msg in errs:
                    self._row(ws, row_num, [counter, tool, "Error", msg],
                              [None, None, _fill_str("FFC7CE")])
                    row_num  += 1
                    counter  += 1
            else:
                self._row(ws, row_num, [counter, tool, "OK", "Extraction completed successfully"],
                          [None, None, _fill_str("C6EFCE")])
                row_num += 1
                counter += 1

        # Any errors from unconfigured tools
        for tool, errs in error_by_tool.items():
            if tool not in estate.configured_tools:
                for msg in errs:
                    self._row(ws, row_num, [counter, tool, "Warning", msg],
                              [None, None, _fill_str("FFFFCC")])
                    row_num += 1
                    counter += 1

        # Estate-level totals
        ws.cell(row=row_num + 1, column=1, value="Totals").font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=row_num + 1, column=2, value=f"{len(estate.configured_tools)} tools configured")
        ws.cell(row=row_num + 1, column=3,
                value=f"{len(estate.summary.extraction_errors)} error(s)")

        self._finalize(ws)
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 70


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _otel_telemetry_hint(component_name: str) -> str:
    """Guess the telemetry type(s) from an OTel component name."""
    n = component_name.lower()
    parts = []
    if any(x in n for x in ("metric", "prometheus", "statsd", "carbon")):
        parts.append("metrics")
    if any(x in n for x in ("trace", "jaeger", "zipkin", "otlp", "tempo")):
        parts.append("traces")
    if any(x in n for x in ("log", "loki", "fluentd", "fluent", "syslog", "filelog")):
        parts.append("logs")
    if "otlp" in n:
        if "metrics" not in parts:
            parts.append("metrics")
        if "traces" not in parts:
            parts.append("traces")
        if "logs" not in parts:
            parts.append("logs")
    return " + ".join(parts) if parts else "unknown"


def _fill_str(hex_color: str) -> str:
    """Return hex string (not PatternFill) for use in _row cell_fills list."""
    return hex_color
